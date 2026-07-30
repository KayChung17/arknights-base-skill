#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Shared data loading helpers for the Arknights base scheduling tools."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, asdict
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable

SKILL_ROOT = Path(__file__).resolve().parents[1]
ASSETS_DIR = SKILL_ROOT / "assets"

OPERATOR_NAME_ALIASES = {
    "阿米娅（医疗）": "阿米娅",
    "阿米娅（近卫）": "阿米娅",
}


@dataclass(frozen=True)
class OwnedOperator:
    name: str
    elite: int = 0
    level: int = 90
    recruited: bool = True
    morale: float | None = None

    def to_dict(self) -> dict:
        return asdict(self)


@lru_cache(maxsize=None)
def load_json(name: str) -> dict:
    path = ASSETS_DIR / name
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


@lru_cache(maxsize=1)
def load_operator_data() -> dict:
    return load_json("operator-skills.json")


@lru_cache(maxsize=1)
def load_mechanics() -> dict:
    return load_json("mechanics.json")


@lru_cache(maxsize=1)
def operator_group_index() -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for item in load_operator_data()["operators"]:
        result.setdefault(item["name"], set()).update(str(group) for group in item.get("groups", []))
    for group, names in (load_json("operator-groups.json").get("groups") or {}).items():
        for name in names:
            result.setdefault(str(name), set()).add(str(group))
    return result


@lru_cache(maxsize=1)
def operator_index() -> dict[str, dict]:
    index = {
        item["name"]: {**item, "groups": list(item.get("groups", []))}
        for item in load_operator_data()["operators"]
    }
    for name, groups in operator_group_index().items():
        if name in index:
            index[name]["groups"] = sorted(groups)
    for alias, canonical in OPERATOR_NAME_ALIASES.items():
        if canonical in index:
            index[alias] = index[canonical]
    return index


def normalize_elite(value: str | int | None) -> int:
    if value is None:
        return 0
    text = str(value).strip().upper()
    if text.startswith("E"):
        text = text[1:]
    try:
        return max(0, min(2, int(float(text))))
    except (TypeError, ValueError):
        return 0


def normalize_bool(value: str | bool | None, default: bool = True) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "已招募"}


def parse_operator_token(token: str) -> OwnedOperator:
    """Parse `name`, `name@E1`, or `name:E2`."""
    token = token.strip()
    if not token:
        raise ValueError("干员名称不能为空")
    for sep in ("@E", "@e", ":E", ":e", "@", ":"):
        if sep in token:
            name, elite = token.rsplit(sep, 1)
            return OwnedOperator(name=name.strip(), elite=normalize_elite(elite))
    return OwnedOperator(name=token, elite=2)


def parse_operator_list(text: str) -> list[OwnedOperator]:
    result = []
    for raw in text.split(","):
        raw = raw.strip()
        if raw:
            result.append(parse_operator_token(raw))
    return result


def read_roster(path: str | Path) -> list[OwnedOperator]:
    """Read TSV, CSV, or XLSX roster files with Chinese or English headers."""
    roster_path = Path(path)
    if not roster_path.exists():
        raise FileNotFoundError(f"干员表不存在: {roster_path}")

    aliases = {
        "name": ["干员名称", "干员名", "name", "operator"],
        "recruited": ["是否已招募", "已招募", "recruited", "owned"],
        "elite": ["精英化等级", "精英等级", "elite", "promotion"],
        "level": ["等级", "level", "operator_level"],
        "morale": ["当前心情", "心情", "morale"],
    }

    def value(row: dict, keys: Iterable[str], default=None):
        for key in keys:
            if key in row and row[key] not in (None, ""):
                return row[key]
        return default

    if roster_path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取XLSX干员表需要安装 openpyxl") from exc
        workbook = load_workbook(roster_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(item).strip() if item is not None else "" for item in next(iterator)]
        except StopIteration:
            workbook.close()
            return []
        raw_rows = list(iterator)
        workbook.close()
        rows = (
            {headers[index]: item[index] if index < len(item) else None for index in range(len(headers))}
            for item in raw_rows
        )
    else:
        text = roster_path.read_text(encoding="utf-8-sig")
        first_line = text.splitlines()[0] if text.splitlines() else ""
        delimiter = "\t" if "\t" in first_line else ","
        rows = csv.DictReader(text.splitlines(), delimiter=delimiter)

    result: list[OwnedOperator] = []
    for row in rows:
        name = str(value(row, aliases["name"], "")).strip()
        if not name:
            continue
        recruited = normalize_bool(value(row, aliases["recruited"], True))
        if not recruited:
            continue
        morale_raw = value(row, aliases["morale"])
        morale = None
        if morale_raw not in (None, ""):
            try:
                morale = float(morale_raw)
            except (TypeError, ValueError):
                morale = None
        result.append(
            OwnedOperator(
                name=name,
                elite=normalize_elite(value(row, aliases["elite"], 0)),
                level=max(1, int(float(value(row, aliases["level"], 1) or 1))),
                recruited=True,
                morale=morale,
            )
        )
    return result


def apply_roster_overrides(
    roster: list[OwnedOperator],
    overrides: dict[str, dict[str, Any]] | None,
) -> list[OwnedOperator]:
    """Apply explicit scenario-only operator state without rewriting the roster file."""
    if not overrides:
        return roster
    result: list[OwnedOperator] = []
    for op in roster:
        override = overrides.get(op.name) or {}
        result.append(OwnedOperator(
            name=op.name,
            elite=normalize_elite(override.get("elite", op.elite)),
            level=max(1, int(override.get("level", op.level) or op.level)),
            recruited=normalize_bool(override.get("recruited", op.recruited), op.recruited),
            morale=(float(override["morale"]) if override.get("morale") is not None else op.morale),
        ))
    return [op for op in result if op.recruited]


def select_available_skills(
    operator_record: dict,
    facility: str,
    elite: int,
    product: str = "",
    level: int = 90,
) -> list[dict]:
    """Choose the highest unlocked variant for each same-named skill."""
    candidates = [
        skill
        for skill in operator_record.get("skills", [])
        if skill.get("facility") == facility
        and int(skill.get("elite", 0)) <= elite
        and int(skill.get("required_level", 1) or 1) <= int(level or 1)
        and (
            not skill.get("products")
            or not product
            or product in skill.get("products", [])
        )
    ]
    grouped: dict[str, list[dict]] = {}
    for item in candidates:
        key = str(item.get("variant_group") or item.get("skill_name", ""))
        grouped.setdefault(key, []).append(item)
    selected = []
    for variants in grouped.values():
        selected.append(max(variants, key=lambda item: (int(item.get("elite", 0)), int(item.get("required_level", 1) or 1))))
    return selected


def display_operator(op: OwnedOperator | dict) -> str:
    if isinstance(op, OwnedOperator):
        return f"{op.name}@E{op.elite}"
    return f"{op.get('name', '')}@E{normalize_elite(op.get('elite', 0))}"
